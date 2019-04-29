import datetime
import io
import os
from datetime import date

from django.conf import settings
from django.core.files.temp import NamedTemporaryFile
from openpyxl import load_workbook
from openpyxl.styles import Font
from import_export import resources


class MyResource(resources.ModelResource):
    excel_template = 'Excel模板/空白.xlsx'
    font = '宋体'

    def before_import(self, dataset, using_transactions, dry_run, **kwargs):
        fields = dict((field.verbose_name, field.name) for field in self._meta.model._meta.fields)
        dataset.headers = [fields[verbose_name.replace('\n', '').replace(' ', '').replace('\t', '')]
                           for verbose_name in dataset.headers]
        super().before_import(dataset, using_transactions, dry_run, **kwargs)

    def before_import_row(self, row, **kwargs):
        replaces = './'
        base_date = date(1900, 1, 1)
        for key, value in row.items():
            if isinstance(value, float) and 49 * 365 < value < (2050 - 1900) * 365:
                value = str(base_date + datetime.timedelta(days=int(value - 2)))
            elif isinstance(value, str):
                if value == '是':
                    value = True
                elif value == '否':
                    value = False
                else:
                    for rep in replaces:
                        value = value.replace(rep, '-')
            row[key] = value
        super(MyResource, self).before_import_row(row, **kwargs)
                            
    def get_export_headers(self):
        headers = super(MyResource, self).get_export_headers()
        model = self._meta.model
        headers_ = []
        for f in headers:
            meta = model._meta
            fields = f.split('__')
            for f1 in fields[:-1]:
                meta = meta.get_field(f1).remote_field.model._meta
            headers_.append(meta.get_field(fields[-1]).verbose_name)
        return headers_

    def to_workbook(self, dataset):
        workbook_name = os.path.join(settings.MEDIA_ROOT, self.excel_template)
        wb = load_workbook(workbook_name)
        page = wb.active
        for row in dataset:
            page.append(row)
        return wb

    def export(self, queryset=None, *args, **kwargs):
        data = super(MyResource, self).export(queryset, *args, **kwargs)
        wb = self.to_workbook(data)
        font = Font(name=self.font)
        for i, row in enumerate(wb.active.iter_rows()):
            if i == 0:
                continue
            for cell in row:
                cell.font = font
        with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)
            output = io.BytesIO(tmp.read())
        wb.close()
        return output.read()
