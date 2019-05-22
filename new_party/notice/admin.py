from common.utils import *
from info.models import Dependency
from info.util import get_visuable_members
from openpyxl import load_workbook
from .util import *


def get_ym(m1, m2):
    if m1 > m2:
        m1, m2 = m2, m1
    today = datetime.datetime.today()
    if today.month <= m2:
        year = today.year
        if today.month <= m1:
            month = m1
        else:
            month = m2
    else:
        year = today.year + 1
        month = m1
    return year, month


class Table:
    excel_template = ''
    beian_template = ''
    row = 3
    fields = []
    verbose_name = '表格'

    @classmethod
    def check(cls, when, queryset):
        """
            检查是否有未及时更新的信息
        """
        return []

    @classmethod
    def get_filters(cls):
        return {}

    @classmethod
    def filter(cls, **kwargs):
        return Member.objects.select_related(None).filter(**kwargs, **cls.get_filters())

    @classmethod
    def get_url(cls):
        filters = cls.get_filters()
        return '&'.join(map(lambda item: '_p_' + item[0] + '=' + str(item[1]), filters.items()))

    @classmethod
    def export_filename(cls):
        return cls.verbose_name + '.xlsx'

    @classmethod
    def export_field(cls):
        return cls.fields

    @classmethod
    def complete_beian(cls, member):
        pass

    @classmethod
    def get_queryset(cls, request):
        qs = queryset(request, cls)
        for q in qs:
            if 'birth_date' in q:
                q['birth_date'] = '-'.join(wrap(q['birth_date']).split('-')[:-1])
        return qs

    @classmethod
    def after_export(cls, sheet, cnt):
        pass

    @classmethod
    def before_export(cls, sheet):
        sheet.cell(1, 1, cls.verbose_name)

    @classmethod
    def export(cls, request):
        work_book = load_workbook(cls.excel_template)
        sheet = work_book.active
        qs = cls.get_queryset(request)
        new_rows = len(qs)
        cls.before_export(sheet)
        insert_rows(sheet, cls.row, new_rows)

        for i, row in enumerate(qs):
            sheet.cell(i + cls.row, 1, i + 1)
            for j, field in enumerate(cls.export_field()):
                value = wrap(row[field])
                sheet.cell(i + cls.row, j + 2, value)
        cls.after_export(sheet, new_rows)
        return to_bytes(work_book)


class FirstTalk(Table):
    excel_template = media('Excel模板/首次组织谈话.xlsx')
    fields = ['branch', 'netid', 'name', 'gender', 'birth_date', 'application_date', 'phone_number',]
              # 'first_talk_end']
    verbose_name = '首次组织谈话'
    phase = verbose_name

    @classmethod
    def get_filters(cls):
        end = datetime.date.today() - datetime.timedelta(days=30)
        return {
            'application_date__gte': end,
            'activist_date__isnull': True,
            'first_talk_date__isnull': True,
            'application_date__isnull': False,
        }

    @classmethod
    def filter(cls, **kwargs):
        return super(FirstTalk, cls).filter(**kwargs)
            # .extra(select={'first_talk_end': 'DATE_ADD(application_date, INTERVAL 1 MONTH)'})


class Activist(Table):
    excel_template = media('Excel模板/入党积极分子.xlsx')
    beian_template = media('Excel模板/材料6：确定入党积极分子等备案.docx')
    fields = ['branch', 'netid', 'name', 'gender', 'birth_date', 'application_date']
    beian_fields = ['branch', 'name', 'gender', 'birth_date', 'id_card_number',
                    'application_date', 'activist_date']
    verbose_name = '%d年%d月可接收入党积极分子' % get_ym(3, 9)
    phase = '入党积极分子'

    @classmethod
    def check(cls, when, queryset):
        deffer = []
        for member in queryset:
            application_date = member['application_date']
            if isinstance(application_date, datetime.datetime):
                application_date = application_date.date()
            month = application_date.month
            interval = (when - application_date).days
            if 2 <= month < 8:
                if when.month > 9 or interval > 4 * 30:
                    remark = member['remarks']
                    if remark is None or (cls.phase + '延迟发展') not in remark:
                        deffer.append(member['netid'])
            else:
                if when.month > 3 or interval > 4 * 30:
                    remark = member['remarks']
                    if remark is None or (cls.phase + '延迟发展') not in remark:
                        deffer.append(member['netid'])
        return deffer

    @classmethod
    def complete_beian(cls, member):
        member.insert(5, '无')
        member.insert(5, '')
        member.insert(8, '团支部推优')
        member.append('')

    @classmethod
    def get_filters(cls):
        year, month = get_ym(3, 9)
        end = datetime.date(year, month, 30)
        try:
            days = Dependency.objects.get(from_1='application_date', to='activist_date').days
            end = end - datetime.timedelta(days=days)
        except Dependency.DoesNotExist:
            pass
        return {
            'application_date__lt': end,
            'activist_date__isnull': True,
            'application_date__isnull': False,
            # first_talk_date__isnull:False,
        }


class KeyDevelop(Table):
    excel_template = media('Excel模板/重点发展对象.xlsx')
    beian_template = media('Excel模板/材料12：确定重点发展对象备案表.docx')
    fields = ['branch', 'netid', 'name', 'gender', 'birth_date', 'application_date', 'activist_date']
    beian_fields = ['branch', 'name', 'gender', 'birth_date', 'id_card_number',
                    'application_date', 'activist_date']
    verbose_name = '%d年%d月可接收重点发展对象' % get_ym(3, 9)
    phase = '重点发展对象'

    @classmethod
    def check(cls, when, queryset):
        deffer = []
        for member in queryset:
            activist_date = member['activist_date']
            if isinstance(activist_date, datetime.datetime):
                activist_date = activist_date.date()
            y, m = when.year, when.month
            if m <= 3:
                m = 9
                y -= 1
            elif 3 < m < 10:
                m = 3
            else:
                m = 9
            when = datetime.date(y, m + 1, 1)
            interval = (when - activist_date).days
            if interval > 365:
                remark = member['remarks']
                if remark is None or (cls.phase + '延迟发展') not in remark:
                    deffer.append(member['netid'])
        return deffer

    @classmethod
    def complete_beian(cls, member):
        member.insert(5, '无')
        member.insert(5, '')
        member.insert(8, '团支部推优')
        member.insert(10, '')

    @classmethod
    def get_filters(cls):
        year, month = get_ym(3, 9)
        end = datetime.date(year, month, 30)
        try:
            days = Dependency.objects.get(from_1='activist_date', to='key_develop_person_date').days
            end = end - datetime.timedelta(days=days)
        except Dependency.DoesNotExist:
            pass
        return {
            'activist_date__lt': end,
            'key_develop_person_date__isnull': True,
            'activist_date__isnull': False,
        }


class LearningClass(Table):
    row = 4
    excel_template = media('Excel模板/材料13：学生入党积极分子党校培训报名汇总表.xlsx')
    fields = ['branch', 'netid', 'name', 'gender', 'group', 'birth_date', 'grade', 'major_in',
              'application_date', 'phone_number']
    verbose_name = '%d年%s季学生入党积极分子党校培训报名汇总表' % (get_ym(4, 10)[0], '春' if get_ym(4, 10)[1] == 4 else '秋')
    phase = ''

    @classmethod
    def export_filename(cls):
        return '材料13：学生入党积极分子党校培训报名汇总表.xlsx'

    @classmethod
    def before_export(cls, sheet):
        super().before_export(sheet)
        try:
            sheet.unmerge_cells('J7:L7')
            sheet.unmerge_cells('J6:L6')
        except:
            pass

    @classmethod
    def after_export(cls, sheet, cnt):
        max_row = sheet.max_row
        sheet.merge_cells('A{row}:M{row}'.format(row=max_row - 2))
        sheet.merge_cells('J{row}:M{row}'.format(row=max_row - 1))
        sheet.row_dimensions[max_row - 1].height = sheet.row_dimensions[max_row].height = 19
        sheet.merge_cells('J{row}:M{row}'.format(row=max_row))

    @classmethod
    def get_filters(cls):
        year, month = get_ym(4, 10)
        end = datetime.date(year, month - 1, 30)
        try:
            days = Dependency.objects.get(from_1='activist_date', to='key_develop_person_date').days
            end = end - datetime.timedelta(days=days)
        except Dependency.DoesNotExist:
            pass
        return {
            'activist_date__lt': end,
            'graduated_party_school_date__isnull': True,
            'activist_date__isnull': False,
        }


class PreMember(Table):
    row = 6
    excel_template = media('Excel模板/材料18：拟吸收预备党员名单汇总审批表.xlsx')
    fields = ['branch', 'netid', 'name', 'birth_date', 'application_date', 'activist_date', 'league_promotion_date',
              'democratic_appraisal_date', 'is_political_check', 'key_develop_person_date',
              'graduated_party_school_date', 'first_branch_conference', 'pro_conversation_date']
    beian_fields = ['branch', 'name', 'gender', 'birth_date', 'id_card_number',
                    'recommenders', 'application_date', 'activist_date',
                    'key_develop_person_date', 'first_branch_conference'
                    ]
    beian_template = media('Excel模板/材料21：接收预备党员备案表.docx')

    verbose_name = '%d年%d月可接收预备党员' % get_ym(6, 12)
    phase = '预备党员'

    @classmethod
    def export_filename(cls):
        return '材料18：拟吸收预备党员名单汇总审批表.xlsx'

    @classmethod
    def export_field(cls):
        return cls.fields[1:]

    @classmethod
    def get_queryset(cls, request):
        qs = super().get_queryset(request)
        qs.sort(key=lambda x: x['netid'])
        return qs

    @classmethod
    def before_export(cls, sheet):
        pass

    @classmethod
    def after_export(cls, sheet, cnt):
        sheet.merge_cells('A{row}:N{row}'.format(row=sheet.max_row))
        sheet.row_dimensions[sheet.max_row - 1].height = 13.5
        sheet.row_dimensions[sheet.max_row].height = 50

    @classmethod
    def complete_beian(cls, member):
        member.insert(5, '无')
        member.insert(5, '')

    @classmethod
    def check(cls, when, queryset):
        deffer = []
        for member in queryset:
            key_develop_person_date = member['key_develop_person_date']
            if isinstance(key_develop_person_date, datetime.datetime):
                key_develop_person_date = key_develop_person_date.date()
            y, m = when.year, when.month
            if m <= 6:
                m = 0
            else:
                m = 6
            when = datetime.date(y, m + 1, 1)
            interval = (when - key_develop_person_date).days
            if interval > 91:
                remark = member['remarks']
                if remark is None or (cls.phase + '延迟发展') not in remark:
                    deffer.append(member['netid'])
        return deffer

    @classmethod
    def get_filters(cls):
        year, month = get_ym(6, 12)
        end = datetime.date(year, month, 30)
        try:
            days = Dependency.objects.get(from_1='key_develop_person_date', to='first_branch_conference').days
            end = end - datetime.timedelta(days=days)
        except Dependency.DoesNotExist:
            pass
        return {
            'key_develop_person_date__lt': end,
            'first_branch_conference__isnull': True,
            'key_develop_person_date__isnull': False,
            'graduated_party_school_date__isnull': False,
        }


class FullMember(Table):
    row = 6
    excel_template = media('Excel模板/材料26：申请转正预备党员名单汇总预审表.xlsx')
    fields = ['branch', 'netid', 'name', 'gender', 'first_branch_conference', 'oach_date',
              'application_fullmember_date']
    verbose_name = '%d年%d月可转正预备党员' % get_ym(6, 12)
    phase = '正式党员'

    @classmethod
    def export_filename(cls):
        return '材料26：申请转正预备党员名单汇总预审表.xlsx'

    @classmethod
    def check(cls, when, queryset):
        deffer = []
        for member in queryset:
            first_branch_conference = member['first_branch_conference']
            if isinstance(first_branch_conference, datetime.datetime):
                first_branch_conference = first_branch_conference.date()
            y, m = when.year, when.month
            if m <= 6:
                m = 0
            else:
                m = 6
            when = datetime.date(y, m + 1, 1)
            interval = (when - first_branch_conference).days
            if interval > 365:
                remark = member['remarks']
                if remark is None or (cls.phase + '延迟发展') not in remark:
                    deffer.append(member['netid'])
        return deffer

    @classmethod
    def get_filters(cls):
        year, month = get_ym(6, 12)
        end = datetime.date(year, month, 30)
        try:
            days = Dependency.objects.get(from_1='first_branch_conference', to='second_branch_conference').days
            end = end - datetime.timedelta(days=days)
        except Dependency.DoesNotExist:
            pass
        return {
            'first_branch_conference__lt': end,
            'second_branch_conference__isnull': True,
            'first_branch_conference__isnull': False,
        }

    @classmethod
    def before_export(cls, sheet):
        pass

    @classmethod
    def after_export(cls, sheet, cnt):
        sheet.merge_cells('A{row}:J{row}'.format(row=sheet.max_row))
        sheet.row_dimensions[sheet.max_row - 1].height = 13.5
        sheet.row_dimensions[sheet.max_row].height = 80
        sheet.cell(4, 1).value = str(sheet.cell(4, 1).value) % cnt


# 提醒更新毕业生组织关系
def get_graduation(request):
    members = get_visuable_members(Member, request.user) \
        .extra(
        where=["MAKEDATE(2000 + netid div 1000000 + years, 120) <= CURDATE()"
               "and (out_type='Z.无' or out_place is null or out_place='')"])
    return members.values('branch', 'netid', 'name', 'out_type', 'out_place')
