from common.utils import *
from info.models import Dependency
from openpyxl import load_workbook
from .util import *


def get_ym(m1, m2):
    if m1 > m2:
        m1, m2 = m2, m1
    today = datetime.datetime.today()
    if today.month <= m2:
        year = today.year
        if today.month <= m1:
            month = m1
        else:
            month = m2
    else:
        year = today.year + 1
        month = m1
    return year, month


class Table:
    excel_template = ''
    beian_template = ''
    row = 3
    fields = []
    verbose_name = '表格'

    @classmethod
    def export_filename(cls):
        return cls.verbose_name + '.xlsx'

    @classmethod
    def export_field(cls):
        return cls.fields

    @classmethod
    def get_queryset(cls, request):
        qs = queryset(request, cls)
        for q in qs:
            if 'birth_date' in q:
                q['birth_date'] = '-'.join(wrap(q['birth_date']).split('-')[:-1])
            q['branch_id'] = q['branch_name']
            del q['branch_name']
        return qs

    @classmethod
    def after_export(cls, sheet, cnt):
        pass

    @classmethod
    def before_export(cls, sheet):
        sheet.cell(1, 1, cls.verbose_name)

    @classmethod
    def export(cls, request):
        work_book = load_workbook(cls.excel_template)
        sheet = work_book.active
        qs = cls.get_queryset(request)
        new_rows = len(qs)
        cls.before_export(sheet)
        insert_rows(sheet, cls.row, new_rows)

        for i, row in enumerate(qs):
            sheet.cell(i + cls.row, 1, i + 1)
            for j, field in enumerate(cls.export_field()):
                value = wrap(row[field])
                sheet.cell(i + cls.row, j + 2, value)
        cls.after_export(sheet, new_rows)
        return to_bytes(work_book)


class FirstTalk(Table):
    excel_template = media('Excel模板/首次组织谈话.xlsx')
    fields = ['branch_id', 'netid', 'name', 'gender', 'birth_date', 'application_date', 'phone_number', 'talk_date_end']
    verbose_name = '首次组织谈话'
    phase = verbose_name

    @staticmethod
    def filter(**kwargs):
        end = datetime.datetime.today() - datetime.timedelta(days=30)
        return Member.objects.select_related(None).filter(**kwargs, activist_date__isnull=True,
                                                          first_talk_date__isnull=True,
                                                          application_date__isnull=False,
                                                          application_date__gte=end).extra(
            select={'talk_date_end': 'DATE_ADD(application_date, INTERVAL 1 MONTH)'})


class Activist(Table):
    excel_template = media('Excel模板/入党积极分子.xlsx')
    fields = ['branch_id', 'netid', 'name', 'gender', 'birth_date', 'application_date']
    verbose_name = '%d年%d月可接收入党积极分子' % get_ym(3, 9)
    phase = '入党积极分子'

    @staticmethod
    def filter(**kwargs):
        year, month = get_ym(3, 9)
        end = datetime.date(year, month, 30)
        try:
            days = Dependency.objects.get(from_1='application_date', to='activist_date').days
            end = end - datetime.timedelta(days=days)
        except Dependency.DoesNotExist:
            pass
        return Member.objects.select_related(None).filter(**kwargs, activist_date__isnull=True,
                                                          application_date__isnull=False,
                                                          first_talk_date__isnull=False,
                                                          application_date__lt=end)


class KeyDevelop(Table):
    excel_template = media('Excel模板/重点发展对象.xlsx')
    fields = ['branch_id', 'netid', 'name', 'gender', 'birth_date', 'application_date', 'activist_date']
    verbose_name = '%d年%d月可接收重点发展对象' % get_ym(3, 9)
    phase = '重点发展对象'

    @staticmethod
    def filter(**kwargs):
        year, month = get_ym(3, 9)
        end = datetime.date(year, month, 30)
        try:
            days = Dependency.objects.get(from_1='activist_date', to='key_develop_person_date').days
            end = end - datetime.timedelta(days=days)
        except Dependency.DoesNotExist:
            pass
        return Member.objects.select_related(None).filter(**kwargs, key_develop_person_date__isnull=True,
                                                          activist_date__isnull=False,
                                                          activist_date__lt=end)


class LearningClass(Table):
    row = 4
    excel_template = media('Excel模板/材料13：学生入党积极分子党校培训报名汇总表.xlsx')
    fields = ['branch_id', 'netid', 'name', 'gender', 'group', 'birth_date', 'grade', 'major_in',
              'application_date', 'phone_number']
    verbose_name = '%d年%s季学生入党积极分子党校培训报名汇总表' % (get_ym(4, 10)[0], '春' if get_ym(4, 10)[1] == 4 else '秋')
    phase = ''

    @classmethod
    def export_filename(cls):
        return '材料13：学生入党积极分子党校培训报名汇总表.xlsx'

    @classmethod
    def before_export(cls, sheet):
        super().before_export(sheet)
        try:
            sheet.unmerge_cells('J7:L7')
            sheet.unmerge_cells('J6:L6')
        except:
            pass

    @classmethod
    def after_export(cls, sheet, cnt):
        max_row = sheet.max_row
        sheet.merge_cells('A{row}:M{row}'.format(row=max_row - 2))
        sheet.merge_cells('J{row}:M{row}'.format(row=max_row - 1))
        sheet.row_dimensions[max_row - 1].height = sheet.row_dimensions[max_row].height = 19
        sheet.merge_cells('J{row}:M{row}'.format(row=max_row))

    @staticmethod
    def filter(**kwargs):
        year, month = get_ym(4, 10)
        end = datetime.date(year, month - 1, 30)
        try:
            days = Dependency.objects.get(from_1='activist_date', to='key_develop_person_date').days
            end = end - datetime.timedelta(days=days)
        except Dependency.DoesNotExist:
            pass
        return Member.objects.select_related(None).filter(**kwargs, graduated_party_school_date__isnull=True,
                                                          activist_date__isnull=False,
                                                          activist_date__lt=end)


class PreMember(Table):
    row = 6
    excel_template = media('Excel模板/材料18：拟吸收预备党员名单汇总审批表.xlsx')
    fields = ['branch_id', 'netid', 'name', 'birth_date', 'application_date', 'activist_date', 'league_promotion_date',
              'democratic_appraisal_date', 'is_political_check', 'key_develop_person_date',
              'graduated_party_school_date', 'first_branch_conference', 'pro_conversation_date']
    beian_fields = ['name', 'gender', 'birth_date', 'id_card_number',
                    'recommenders', 'application_date', 'activist_date',
                    'key_develop_person_date', 'first_branch_conference'
                    ]
    beian_template = media('Excel模板/材料21：接收预备党员备案表.docx')

    verbose_name = '%d年%d月可接收预备党员' % get_ym(6, 12)
    phase = '预备党员'

    @classmethod
    def export_filename(cls):
        return '材料18：拟吸收预备党员名单汇总审批表.xlsx'

    @classmethod
    def export_field(cls):
        return cls.fields[1:]

    @classmethod
    def get_queryset(cls, request):
        qs = super().get_queryset(request)
        qs.sort(key=lambda x: x['netid'])
        return qs

    @classmethod
    def before_export(cls, sheet):
        pass

    @classmethod
    def after_export(cls, sheet, cnt):
        sheet.merge_cells('A{row}:N{row}'.format(row=sheet.max_row))
        sheet.row_dimensions[sheet.max_row - 1].height = 13.5
        sheet.row_dimensions[sheet.max_row].height = 50

    @staticmethod
    def filter(**kwargs):
        year, month = get_ym(6, 12)
        end = datetime.date(year, month, 30)
        try:
            days = Dependency.objects.get(from_1='key_develop_person_date', to='first_branch_conference').days
            end = end - datetime.timedelta(days=days)
        except Dependency.DoesNotExist:
            pass
        return Member.objects.select_related(None).filter(**kwargs, first_branch_conference__isnull=True,
                                                          key_develop_person_date__isnull=False,
                                                          key_develop_person_date__lt=end)


class FullMember(Table):
    row = 6
    excel_template = media('Excel模板/材料26：申请转正预备党员名单汇总预审表.xlsx')
    fields = ['branch_id', 'netid', 'name', 'gender', 'first_branch_conference', 'oach_date',
              'application_fullmember_date']
    verbose_name = '%d年%d月可转正预备党员' % get_ym(6, 12)
    phase = '正式党员'

    @classmethod
    def export_filename(cls):
        return '材料26：申请转正预备党员名单汇总预审表.xlsx'

    @staticmethod
    def filter(**kwargs):
        year, month = get_ym(6, 12)
        end = datetime.date(year, month, 30)
        try:
            days = Dependency.objects.get(from_1='first_branch_conference', to='second_branch_conference').days
            end = end - datetime.timedelta(days=days)
        except Dependency.DoesNotExist:
            pass
        return Member.objects.select_related(None).filter(**kwargs, second_branch_conference__isnull=True,
                                                          first_branch_conference__isnull=False,
                                                          first_branch_conference__lt=end)

    @classmethod
    def before_export(cls, sheet):
        pass

    @classmethod
    def get_queryset(cls, request):
        qs = queryset(request, cls)
        for q in qs:
            q['branch_id'] = q['branch_name']
        return qs

    @classmethod
    def after_export(cls, sheet, cnt):
        sheet.merge_cells('A{row}:J{row}'.format(row=sheet.max_row))
        sheet.row_dimensions[sheet.max_row - 1].height = 13.5
        sheet.row_dimensions[sheet.max_row].height = 80
        sheet.cell(4, 1).value = str(sheet.cell(4, 1).value) % cnt
