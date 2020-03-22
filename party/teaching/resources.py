import os

from django.conf import settings
from openpyxl import load_workbook
from openpyxl.styles import Alignment

from common import resources
from common.utils import set_align
from .models import *
from info.models import Member, Branch


class ActivityResource(resources.MyResource):
    excel_template = 'Excel模板/会议和活动.xlsx'

    class Meta:
        model = Activity
        skip_unchanged = True
        import_id_fields = ('name', 'date')
        fields = ['name', 'date', 'branch', 'credit']

    def export_resource(self, obj: Activity):
        return [
            obj.name,
            obj.date.strftime('%Y-%m-%d %H:%M'),
            '\n'.join([b.branch_name for b in obj.branch.all()]),
            obj.credit
        ]

    def to_workbook(self, dataset):
        wb = super(ActivityResource, self).to_workbook(dataset)
        for i, row in enumerate(wb.active.iter_rows()):
            if i == 0:
                continue
            max_row = 1
            default = 14.25
            for cell in row:
                tmp = str(cell.value).count('\n') + 1
                if tmp > max_row:
                    max_row = tmp
            if max_row != 1:
                wb.active.row_dimensions[i + 1].height = max_row * default
        set_align(wb, 'center', 'center')
        return wb


def _group_to_dict(dataset):
    dataset1 = dict()
    for netid, activity_id, credit in dataset:
        m = Member.objects.get(netid=netid)
        activity = Activity.objects.get(id=activity_id)
        branch_id = m.branch_id
        dataset1.setdefault(branch_id, dict())
        dataset1[branch_id].setdefault(netid, [])
        dataset1[branch_id][netid].append((activity.name, activity.date.strftime("%Y-%m-%d %H:%M"), credit))
    return dataset1


class CreditResource(resources.MyResource):
    excel_template = 'Excel模板/学时统计.xlsx'
    import_excel = 'Excel模板/学时导入.xlsx'

    class Meta:
        model = TakePartIn
        skip_unchanged = True
        import_id_fields = ('member', 'activity')
        fields = ('member', 'activity', 'credit')

    def to_workbook(self, dataset):
        dataset = _group_to_dict(dataset)
        workbook = load_workbook(os.path.join(settings.MEDIA_ROOT, self.excel_template))
        sheet = workbook.active
        alignment = Alignment(horizontal='center', vertical='center')
        i = 2
        for branch in sorted(dataset.keys()):
            row = i
            sheet.cell(i, 1, Branch.objects.get(id=branch).branch_name).alignment = alignment
            member_takes = dataset[branch]
            for netid in sorted(member_takes.keys(),
                                key=lambda x: sum(map(lambda x: x[2], member_takes[x])), reverse=True):
                member = Member.objects.get(netid=netid)
                sheet.cell(i, 2, member.netid).alignment = alignment
                sheet.cell(i, 3, member.name).alignment = alignment
                sheet.cell(i, 4, member.get_identity()).alignment = alignment
                takes = member_takes[netid]
                j = i
                for take in takes:
                    sheet.cell(i, 6, take[0]).alignment = alignment
                    sheet.cell(i, 7, take[1]).alignment = alignment
                    sheet.cell(i, 8, take[2]).alignment = alignment
                    i += 1
                sheet.cell(j, 5, '=SUM(H%d:H%d)' % (j, i - 1)).alignment = alignment
                for column in 'BCDE':
                    sheet.merge_cells('%s%d:%s%d' % (column, j, column, i - 1))
            sheet.merge_cells('A%d:A%d' % (row, i - 1))
        return workbook

    def before_import(self, dataset, using_transactions, dry_run, **kwargs):
        header = []
        for h in dataset.headers:
            if '（' in h:
                h = h[: h.find('（')]
            header.append(h)
        dataset.headers = header
        super(CreditResource, self).before_import(dataset, using_transactions, dry_run, **kwargs)


class AskForLeaveResource(resources.MyResource):
    excel_template = 'Excel模板/请假、缺席统计.xlsx'

    class Meta:
        model = AskForLeave
        skip_unchanged = True
        import_id_fields = ('member', 'activity')
        fields = ('member', 'activity', 'status')

    def to_workbook(self, dataset):
        dataset = _group_to_dict(dataset)
        workbook = load_workbook(os.path.join(settings.MEDIA_ROOT, self.excel_template))
        sheet = workbook.active
        alignment = Alignment(horizontal='center', vertical='center')
        status = dict(AskForLeave._meta.get_field('status').choices)
        i = 2
        for branch in sorted(dataset.keys()):
            row = i
            sheet.cell(i, 1, Branch.objects.get(id=branch).branch_name).alignment = alignment
            member_takes = dataset[branch]
            for netid in member_takes:
                member = Member.objects.get(netid=netid)
                sheet.cell(i, 2, member.netid).alignment = alignment
                sheet.cell(i, 3, member.name).alignment = alignment
                sheet.cell(i, 4, member.get_identity()).alignment = alignment
                takes = member_takes[netid]
                j = i
                for take in takes:
                    sheet.cell(i, 5, take[0]).alignment = alignment
                    sheet.cell(i, 6, take[1]).alignment = alignment
                    sheet.cell(i, 7, status[take[2]]).alignment = alignment
                    i += 1
                for column in 'BCD':
                    sheet.merge_cells('%s%d:%s%d' % (column, j, column, i - 1))
            sheet.merge_cells('A%d:A%d' % (row, i - 1))
        return workbook
