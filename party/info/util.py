import datetime
import io
import os
from collections import OrderedDict

from django.conf import settings
from django.contrib import messages
from django.core.files.temp import NamedTemporaryFile
from django.db.models import Q
from django.http import HttpResponse
from django.utils.encoding import escape_uri_path
from openpyxl import load_workbook
from openpyxl.styles import Border, Side

from common.base import get_old, get_chinese
from common.rules import *
from common.utils import set_font, set_align, to_bytes
from info.models import Member, Branch, Dependency


def get_end_time(days):
    now = datetime.datetime.now()
    delta = datetime.timedelta(days=days)
    return now - delta, now.month


def group_by_branch(appers):
    groups = dict()
    for apper in appers:
        groups.setdefault(apper['branch_id'], [])
        groups[apper['branch_id']].append(apper)
    return groups


def get_visuable_members(model, user):
    qs = model.objects
    if is_branch_manager(user):  # 支书
        member = user.member
        if member is not None:
            return qs.filter(branch_id=member['branch_id'])
    elif is_member(user):
        member = user.member
        if member is not None:
            return qs.filter(netid=member['netid']) or qs.none()
    elif is_school_manager(user):
        school = user.school_id
        return qs.filter(branch__school_id=school)
    elif user.is_superuser:
        return qs.all()
    return qs.none()


def get_visual_branch(user):
    if is_school_manager(user):
        school = user.school_id
        return Branch.objects.filter(school_id=school)
    elif is_branch_manager(user) or is_member(user):  # 支书
        member = user.member
        if member is not None:
            return Branch.objects.filter(id=member['branch_id'])
    elif user.is_superuser:
        return Branch.objects.all()


def check_date_dep(obj, old):
    errors = []
    for dep in Dependency.objects.filter(Q(scope=0) | Q(scope=1 + int(not obj.is_sysu))):
        from_ = getattr(obj, dep.from_1)
        if isinstance(from_, datetime.datetime):
            from_ = from_.date()
        to = getattr(obj, dep.to)
        if isinstance(to, datetime.datetime):
            to = to.date()
        from_2 = None if old is None else getattr(old, dep.from_1)
        to2 = None if old is None else getattr(old, dep.to)
        if (from_ != from_2 or to != to2) and from_ and to:
            delta = to - from_
            if delta.days < dep.days:
                errors.append((dep.from_1, dep.to, delta.days,
                               dep.days_mapping[dep.days]))
    return errors


def check_first_talk_date(obj, old):
    if obj.first_talk_date and obj.application_date:
        if old is None or (obj.application_date != old.application_date or
                           obj.first_talk_date != old.first_talk_date):
            days = (obj.first_talk_date - obj.application_date).days
            return days < 31
    return True


def get_members(branch, names):
    res = []
    for name in names:
        try:
            res.append(Member.objects.filter(branch_id=branch, name=name).first())
        except Member.DoesNotExist:
            pass
    return res


def check_fields(obj, error):
    old = get_old(obj)

    errors = check_date_dep(obj, old)
    for e in errors:
        error.append("%s到%s需要%s，而%s只用了%d天。"
                     % (Member._meta.get_field(e[0]).verbose_name.strip('时间'),
               Member._meta.get_field(e[1]).verbose_name.strip('时间'),
               e[3], obj, e[2]))
        return False
    # 检查首次组织谈话时间
    if obj.is_sysu and not check_first_talk_date(obj, old):
        error.append('未在一个月内完成首次组织谈话。')
        return False
    # 检查入党介绍人
    if obj.is_sysu and (old is None or obj.recommenders != old.recommenders):
        for m in get_members(obj.branch_id, get_chinese(str(obj.recommenders))):
            if not m.is_real_party_member():
                error.append('入党介绍人%s不是正式党员。' % m.name)
                return False
    return True


def when_dangxiao(member):
    today = datetime.datetime.today()
    if member.key_develop_person_date:
        return 'F' if 4 < today.month < 10 else 'S'
    else:
        date_key_person = member.activist_date + datetime.timedelta(days=365)
        if date_key_person.year != today.year:
            return 'N'
        return 'F' if 4 < date_key_person.month < 10 else 'S'


def export_statistics(request):
    members = get_visuable_members(Member, request.user)
    columns = ['支部人数', '党员比例',
               '正式党员数', '预备党员数', '积极分子比例', '入党积极分子数',
               '提交入党申请书人数', '春季党训班可报送人数', '秋季党训班可报送人数'
               ]
    statistics = dict()
    for member in members:
        branch = member.branch.branch_name
        grade = member.grade()
        statistics.setdefault(branch, dict())
        statistics[branch].setdefault(grade, OrderedDict([(c, 0) for c in columns]))
        row = statistics[branch][grade]
        row['支部人数'] += 1
        if member.is_real_party_member():
            row['正式党员数'] += 1
        elif member.is_pre_party_member():
            row['预备党员数'] += 1
        elif member.activist_date:
            row['入党积极分子数'] += 1
            if not member.graduated_party_school_date:
                when = when_dangxiao(member)
                if when == 'F':
                    row['秋季党训班可报送人数'] += 1
                elif when == 'S':
                    row['春季党训班可报送人数'] += 1
        elif member.application_date:
            row['提交入党申请书人数'] += 1
    wb = load_workbook(os.path.join(settings.MEDIA_ROOT, 'Excel模板/党员数据表.xlsx'))
    sheet = wb.active
    bottom_border = Border(bottom=Side(style='thin'), right=Side(style='thin'))
    right_border = Border(right=Side(style='thin'))
    begin_row = sheet.max_row
    i = begin_row + 1
    for branch, branch_data in statistics.items():
        j = i
        for grade, grade_data in branch_data.items():
            cnt = grade_data['支部人数']
            if cnt:
                rate = (grade_data['正式党员数'] + grade_data['预备党员数']) / cnt
                grade_data['党员比例'] = "%.1f%%" % (rate * 100)
                rate = grade_data['入党积极分子数'] / cnt
                grade_data['积极分子比例'] = "%.1f%%" % (rate * 100)
            sheet.append([branch, grade] + list(grade_data.values()))
            for cell in sheet[i]:
                cell.border = right_border
            i += 1
        for cell in sheet[i - 1]:
            cell.border = bottom_border
        sheet.merge_cells('A%d:A%d' % (j, i - 1))
    set_font(wb, '宋体', begin_row)
    set_align(wb, 'center', 'center', begin_row)
    data = to_bytes(wb)
    wb.close()
    content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response = HttpResponse(data, content_type=content_type)
    response['Content-Disposition'] = "attachment; filename*=utf-8''{}".format(escape_uri_path(
        '党员数据表.xlsx'
    ))
    return response


def get_detail_chart(request):
    m = request.user.member
    if m is None and not is_school_admin(request.user):
        return None
    try:
        branch = int(request.GET.get('branch') or request.path.split('/')[-3])
    except ValueError:
        return None

    important_dates = ('netid', 'application_date', 'activist_date',
                       'key_develop_person_date', 'first_branch_conference',
                       'second_branch_conference')
    if is_school_admin(request.user):
        objects = Member.objects.filter(branch_id=branch).values(*important_dates)
        scope = Branch.objects.get(id=branch).branch_name
    else:
        scope = '党支部'
        objects = Member.objects.filter(branch_id=branch).values(*important_dates)
    if not objects.exists():
        return None
    return make_chart(objects, scope)


def make_chart(objects, scope):
    my_charts = {
        'fenbu': {
            'title': scope + '成员构成'
        }
    }
    dates = OrderedDict([
        ('second_branch_conference', '正式党员'),
        ('first_branch_conference', '预备党员'),
        ('key_develop_person_date', '重点发展对象'),
        ('activist_date', '入党积极分子'),
        ('application_date', '入党申请人')
    ])

    fenbu = dict()
    for obj in objects:
        grade = '20' + str(obj['netid'])[:2]
        fenbu.setdefault(grade, OrderedDict(
            [(k, 0) for k in dates.values()]))
        for d in dates:
            if obj[d]:
                fenbu[grade][dates[d]] += 1
                break
    grades = list(sorted(fenbu.keys()))
    option = {
        'tooltip': {
            'trigger': 'axis',
            'axisPointer': {
                'type': 'shadow'
            }
        },
        'toolbox': {
            'feature': {
                'saveAsImage': {'show': True}
            }
        },
        'legend': {
            'data': list(dates.values())
        },
        'grid': {
            'left': '3%',
            'right': '4%',
            'bottom': '3%',
            'containLabel': True
        },
        'xAxis': {
            'type': 'value'
        },
        'yAxis': {
            'type': 'category',
            'data': [g + '级' for g in grades]
        },
        'series': [
            {
                'name': name,
                'type': 'bar',
                'stack': '总量',
                'label': {
                    'normal': {
                        'show': False,
                        'position': 'insideRight'
                    }
                },
                'data': [fenbu[g][name] for g in grades]
            } for name in dates.values()
        ]
    }
    my_charts['fenbu']['option'] = option
    return my_charts


def get_list_chart(request):
    important_dates = ('netid', 'application_date', 'activist_date',
                       'key_develop_person_date', 'first_branch_conference',
                       'second_branch_conference')
    if is_school_admin(request.user):
        objects = Member.objects.all().values(*important_dates)
        scope = '数据科学与计算机学院'
    else:
        m = request.user.member
        if m is None:
            return None
        scope = '党支部'
        objects = Member.objects.filter(branch_id=m['branch_id']).values(*important_dates)
    if not objects.exists():
        return None
    return make_chart(objects, scope)
